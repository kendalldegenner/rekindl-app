import streamlit as st
import zipfile
import json
import io
import re
import os
import csv
from collections import Counter
from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────
st.set_page_config(
    page_title="Rekindl — Client Onboarding Tool",
    page_icon="🔥",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ─────────────────────────────────────────
# STYLING
# ─────────────────────────────────────────
st.markdown("""
<style>
    .rekindl-header {
        background: linear-gradient(135deg, #1A1A2E 0%, #0F3460 100%);
        padding: 2rem;
        border-radius: 12px;
        margin-bottom: 2rem;
        text-align: center;
    }
    .rekindl-header h1 { color: #ffffff; font-size: 2.5rem; margin: 0; }
    .rekindl-header p  { color: #aaaacc; font-size: 1rem; margin: 0.5rem 0 0; }
    .metric-card {
        background: #f8f9fa;
        border-left: 4px solid #E94560;
        border-radius: 8px;
        padding: 1rem 1.2rem;
        margin-bottom: 0.5rem;
    }
    .metric-card h3 { margin: 0; color: #1A1A2E; font-size: 1.8rem; }
    .metric-card p  { margin: 0; color: #666; font-size: 0.85rem; }
    .voice-quote {
        background: #f0f4ff;
        border-left: 4px solid #0F3460;
        border-radius: 6px;
        padding: 0.8rem 1rem;
        font-style: italic;
        color: #333;
        margin: 0.5rem 0;
    }
    .tag-chip {
        display: inline-block;
        background: #E94560;
        color: white;
        border-radius: 20px;
        padding: 2px 10px;
        font-size: 0.78rem;
        margin: 2px;
    }
    .stDownloadButton > button {
        background-color: #E94560 !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        font-weight: 600 !important;
    }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="rekindl-header">
    <h1>🔥 Rekindl</h1>
    <p>AI-Powered Lead Reactivation — Client Onboarding Tool</p>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────

def fix_encoding(text):
    try:
        return text.encode('latin1').decode('utf-8')
    except Exception:
        return text

def clean_phone(raw):
    if not raw:
        return None
    digits = re.sub(r'[^\d]', '', str(raw))
    if len(digits) == 10:
        return '+1' + digits
    if len(digits) == 11 and digits[0] == '1':
        return '+' + digits
    return None

def extract_phones_from_text(text):
    patterns = [
        r'\b1[-.\s]?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}\b',
        r'\b\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}\b',
        r'\b\d{3}[-.\s]\d{3}[-.\s]\d{4}\b',
    ]
    found = []
    for p in patterns:
        found += re.findall(p, text)
    return found

NOW = datetime.now(timezone.utc)

# ─────────────────────────────────────────
# STEP 1 — LOAD ZIP
# ─────────────────────────────────────────

def load_conversations_from_zip(zip_bytes, sender_name):
    """Extract all conversations from a Facebook Messenger zip export."""
    conversations = []
    sender_lower = sender_name.lower().strip()

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        message_files = [
            f for f in zf.namelist()
            if re.search(r'messages/inbox/[^/]+/message_\d+\.json', f)
        ]

        folder_map = {}
        for mf in message_files:
            parts = mf.split('/')
            folder = parts[parts.index('inbox') + 1]
            folder_map.setdefault(folder, []).append(mf)

        for folder, files in folder_map.items():
            all_messages = []
            participants = []
            for f in sorted(files):
                with zf.open(f) as fh:
                    data = json.load(fh)
                    if not participants:
                        participants = [fix_encoding(p.get('name', '')) for p in data.get('participants', [])]
                    msgs = data.get('messages', [])
                    for m in msgs:
                        sender = fix_encoding(m.get('sender_name', ''))
                        content = fix_encoding(m.get('content', ''))
                        ts = m.get('timestamp_ms', 0) / 1000
                        all_messages.append({
                            'sender': sender,
                            'content': content,
                            'timestamp': ts
                        })

            if not all_messages:
                continue

            # Determine the other participant's name
            other_names = [p for p in participants if p.lower() != sender_lower]
            contact_name = other_names[0] if other_names else folder

            # Split messages by sender
            my_messages = [m for m in all_messages if m['sender'].lower() == sender_lower]
            all_messages_sorted = sorted(all_messages, key=lambda x: x['timestamp'])
            last_ts = all_messages_sorted[-1]['timestamp'] if all_messages_sorted else 0
            last_msg = all_messages_sorted[-1]['content'][:120] if all_messages_sorted else ''

            days_since = int((NOW.timestamp() - last_ts) / 86400) if last_ts else 9999

            # Extract phones from all content
            all_text = ' '.join(m['content'] for m in all_messages)
            raw_phones = extract_phones_from_text(all_text)
            phone = None
            for rp in raw_phones:
                p = clean_phone(rp)
                if p:
                    phone = p
                    break

            conversations.append({
                'folder': folder,
                'contact_name': contact_name,
                'participants': participants,
                'my_messages': [m['content'] for m in my_messages],
                'all_messages': all_messages_sorted,
                'total_messages': len(all_messages),
                'my_message_count': len(my_messages),
                'last_timestamp': last_ts,
                'last_message': last_msg,
                'days_since': days_since,
                'phone': phone,
            })

    return conversations

# ─────────────────────────────────────────
# STEP 2 — BRAND VOICE ANALYSIS
# ─────────────────────────────────────────

TONE_MARKERS = {
    'Warm & Friendly':    ['hey', 'hi', 'hello', 'how are you', 'hope you', 'good to', 'great to', 'nice to', 'lovely', 'wonderful', 'awesome', 'amazing'],
    'Direct & Action':    ["let's", "let me", 'call me', 'reach out', 'contact', 'come in', 'stop by', 'book', 'schedule', 'set up', 'get you', "i'd be happy"],
    'Community-First':    ['family', 'community', 'together', 'our', 'we', 'team', 'everyone', 'folks', 'people', 'neighbour', 'neighbor'],
    'Casual & Relaxed':   ['yup', 'yep', 'yeah', 'nope', 'gonna', 'wanna', 'kinda', 'sorta', 'tbh', 'lol', 'haha', 'omg', 'no worries'],
    'Encouraging':        ['you got this', "you're amazing", 'proud of you', 'great job', 'well done', 'keep it up', "you're doing", 'love that', 'so happy for'],
    'Professional':       ['please', 'thank you', 'sincerely', 'regarding', 'as per', 'further to', 'following up', 'kindly', 'at your earliest'],
}

OPENER_PATTERNS = ['hey', 'hi', 'hello', 'good morning', 'good afternoon', 'good evening', 'morning', 'afternoon']

def analyse_voice(my_messages):
    if not my_messages:
        return {}

    all_text = ' '.join(my_messages).lower()
    words = re.findall(r"\b[a-z']+\b", all_text)
    word_freq = Counter(words)

    stopwords = {'the','a','an','and','or','but','in','on','at','to','for','of','with','is','it',
                 'i','you','he','she','we','they','this','that','these','those','was','are','be',
                 'have','has','had','do','did','will','would','could','should','my','your','our',
                 'me','him','her','us','them','if','so','as','by','from','up','out','not','no',
                 'can','just','get','got','been','also','what','when','where','who','how','about'}
    top_words = [(w, c) for w, c in word_freq.most_common(60) if w not in stopwords and len(w) > 2][:20]

    bigrams = Counter()
    for msg in my_messages:
        ws = re.findall(r"\b[a-z']+\b", msg.lower())
        for i in range(len(ws)-1):
            bigrams[(ws[i], ws[i+1])] += 1
    top_bigrams = [(' '.join(bg), c) for bg, c in bigrams.most_common(15) if c > 1]

    trigrams = Counter()
    for msg in my_messages:
        ws = re.findall(r"\b[a-z']+\b", msg.lower())
        for i in range(len(ws)-2):
            trigrams[(ws[i], ws[i+1], ws[i+2])] += 1
    top_trigrams = [(' '.join(tg), c) for tg, c in trigrams.most_common(10) if c > 2]

    tone_scores = {}
    for tone, markers in TONE_MARKERS.items():
        score = sum(all_text.count(m) for m in markers)
        tone_scores[tone] = score

    openers = Counter()
    for msg in my_messages:
        first = msg.strip().lower()
        for op in OPENER_PATTERNS:
            if first.startswith(op):
                openers[op] += 1
                break

    emoji_msgs = [m for m in my_messages if any(ord(c) > 8000 for c in m)]
    emoji_pct = round(len(emoji_msgs) / len(my_messages) * 100, 1) if my_messages else 0

    end_punct = Counter()
    for msg in my_messages:
        stripped = msg.strip()
        if stripped:
            last = stripped[-1]
            if last in '.!?':
                end_punct[last] += 1
            else:
                end_punct['none'] += 1
    no_punct_pct = round(end_punct.get('none', 0) / len(my_messages) * 100, 1) if my_messages else 0

    avg_len = round(sum(len(m.split()) for m in my_messages) / len(my_messages), 1) if my_messages else 0

    return {
        'total_messages': len(my_messages),
        'avg_message_length': avg_len,
        'top_words': top_words,
        'top_bigrams': top_bigrams,
        'top_trigrams': top_trigrams,
        'tone_scores': tone_scores,
        'openers': openers.most_common(5),
        'emoji_pct': emoji_pct,
        'no_punct_pct': no_punct_pct,
    }

# ─────────────────────────────────────────
# STEP 3 — OPPORTUNITY CLASSIFICATION
# ─────────────────────────────────────────

SALE_SIGNALS     = ['approved', 'approval', 'delivered', 'delivery', 'picked up', 'financing', 'financed',
                    'deposit', 'down payment', 'signed', 'plates', 'registered', 'congratulations', 'congrats',
                    'welcome to the family', 'keys', 'drove away', 'just bought', 'just got', 'closed']
INTEREST_SIGNALS = ['interested', 'how much', 'what\'s the price', 'price', 'payment', 'monthly',
                    'trade', 'trade-in', 'upgrade', 'looking for', 'need a', 'want a', 'shopping',
                    'test drive', 'come in', 'availability', 'available', 'in stock', 'lease', 'finance',
                    'sounds good', 'maybe', 'possibly', 'thinking about', 'considering', 'might be']
DECLINE_SIGNALS  = ['not interested', 'no thanks', 'not right now', 'already bought', 'already have',
                    'went somewhere else', 'bought elsewhere', 'got one', 'never mind', 'cancel', 'stop',
                    'remove me', 'unsubscribe', 'leave me alone', 'do not contact']

def classify_conversation(conv):
    all_text = ' '.join(m['content'] for m in conv['all_messages']).lower()
    my_text  = ' '.join(conv['my_messages']).lower()

    sale_count     = sum(1 for s in SALE_SIGNALS     if s in all_text)
    interest_count = sum(1 for s in INTEREST_SIGNALS if s in all_text)
    decline_count  = sum(1 for s in DECLINE_SIGNALS  if s in all_text)

    days = conv['days_since']
    my_last = conv['my_messages'][-1] if conv['my_messages'] else ''
    all_msgs = conv['all_messages']
    last_sender = all_msgs[-1]['sender'].lower() if all_msgs else ''

    # Check if they (not Amy) sent the last message
    amy_sent_last = last_sender == 'amy gauthier' or conv['my_message_count'] > 0 and all_msgs and all_msgs[-1]['sender'].lower() in [p.lower() for p in conv['participants'] if p.lower() != conv['contact_name'].lower()]

    evidence = []
    if sale_count:     evidence.append(f"Sale confirmed ({sale_count} signals)")
    if interest_count: evidence.append(f"Interest signals ({interest_count})")
    if decline_count:  evidence.append(f"Decline signals ({decline_count})")
    evidence_str = '. '.join(evidence) + f'. {days} days ago.' if evidence else f'Last contact {days} days ago.'

    # Classification logic
    if decline_count >= 2:
        return 'Not Interested', 1, 'Low priority — revisit in 12 months.'
    if sale_count >= 2 and days < 180:
        return 'Recent Customer – Stay Warm', 8, 'Recent buyer — ask for referrals and check satisfaction.'
    if sale_count >= 2 and days >= 180:
        p = 9; cat = 'Past Customer – Trade Up Ready'
        return cat, p, "High priority — they've already bought from you. Lead with trade-up value and loyalty."
    if sale_count == 1 and days >= 365:
        return 'Past Customer – Check In', 7, "Past buyer — check in and see if they're ready to upgrade."
    # Last message was from THEM (unanswered)
    if all_msgs and all_msgs[-1]['sender'].lower() != my_text[:5].lower():
        if days < 30:
            return "Amy Needs to Reply", 8, "They reached out recently — reply now."
        elif days < 90:
            return "Amy Needs to Reply", 7, "They reached out — reply before the trail goes cold."
    if interest_count >= 3 and days < 90:
        return 'Hot Lead – Needs Follow-Up', 8, 'Strong interest signals — follow up this week.'
    if interest_count >= 2 and days < 180:
        return 'Hot Lead – Went Cold', 7, 'Showed real interest but went quiet — re-engage with a personal message.'
    if interest_count >= 1 and days < 365:
        return 'Warm Lead – Engaged', 6, 'Some interest — keep warm with a value-first message.'
    if conv['my_message_count'] > 0 and days < 60:
        return 'Unanswered Outreach – Recent', 4, 'You reached out recently with no reply — try one more time.'
    if conv['my_message_count'] > 0 and days < 730:
        return 'Unanswered Outreach – Old', 3, 'Old unanswered outreach — rekindle with fresh approach.'
    if days < 90:
        return 'Lead – Needs Follow-Up', 5, 'Recent conversation with potential — follow up now.'
    return 'Cold Lead – No Reply', 2, 'No engagement — try a completely fresh approach.'

CATEGORY_ORDER = [
    'Past Customer – Trade Up Ready',
    'Amy Needs to Reply',
    'Recent Customer – Stay Warm',
    'Past Customer – Check In',
    'Hot Lead – Needs Follow-Up',
    'Hot Lead – Went Cold',
    'Lead – Needs Follow-Up',
    'Warm Lead – Engaged',
    'Unanswered Outreach – Recent',
    'Cold Lead – No Reply',
    'Unanswered Outreach – Old',
    'Not Interested',
]

CATEGORY_COLORS = {
    'Past Customer – Trade Up Ready': 'FF4757',
    'Amy Needs to Reply':             'FF6B35',
    'Recent Customer – Stay Warm':    'FF9F43',
    'Past Customer – Check In':       'FFC312',
    'Hot Lead – Needs Follow-Up':     'C4E538',
    'Hot Lead – Went Cold':           '7BED9F',
    'Lead – Needs Follow-Up':         '70A1FF',
    'Warm Lead – Engaged':            '5352ED',
    'Unanswered Outreach – Recent':   'ECCC68',
    'Cold Lead – No Reply':           'A4B0BE',
    'Unanswered Outreach – Old':      'DFE6E9',
    'Not Interested':                 'B2BEC3',
}

def classify_all(conversations):
    results = []
    for conv in conversations:
        if conv['my_message_count'] == 0 and len(conv['all_messages']) < 2:
            continue
        cat, priority, suggested = classify_conversation(conv)
        results.append({
            'contact_name': conv['contact_name'],
            'category': cat,
            'priority': priority,
            'days_since': conv['days_since'],
            'phone': conv['phone'],
            'last_message': conv['last_message'],
            'suggested': suggested,
            'folder': conv['folder'],
            'my_messages': len(conv['my_messages']),
            'total_messages': len(conv['all_messages']),
        })

    results.sort(key=lambda x: (
        CATEGORY_ORDER.index(x['category']) if x['category'] in CATEGORY_ORDER else 99,
        -x['priority'],
        x['days_since']
    ))
    return results

# ─────────────────────────────────────────
# STEP 4 — GENERATE EXCEL
# ─────────────────────────────────────────

def build_excel(classified, sender_name):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    thin = Side(style='thin', color='DDDDDD')
    def cell_border():
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def header_style(ws, row, cols, fill_color='1A1A2E', font_color='FFFFFF', height=28):
        for col_i, val in enumerate(cols, 1):
            c = ws.cell(row=row, column=col_i, value=val)
            c.fill = PatternFill('solid', fgColor=fill_color)
            c.font = Font(bold=True, color=font_color, size=10, name='Arial')
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = cell_border()
        ws.row_dimensions[row].height = height

    # ── Tab 1: All Opportunities ──
    ws1 = wb.create_sheet("📋 All Opportunities")
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = 'A3'

    ws1.merge_cells('A1:H1')
    title_cell = ws1['A1']
    title_cell.value = f"Rekindl — Sales Opportunity Tracker | {sender_name} | {datetime.now().strftime('%B %d, %Y')}"
    title_cell.font = Font(bold=True, size=13, color='FFFFFF', name='Arial')
    title_cell.fill = PatternFill('solid', fgColor='1A1A2E')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 36

    header_style(ws1, 2, ['Contact Name', 'Category', 'Priority', 'Days Since Contact', 'Phone', 'Last Message Preview', 'Suggested Outreach', 'Folder'])
    col_widths = [22, 28, 10, 18, 16, 40, 45, 28]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    for r_idx, row in enumerate(classified, 3):
        color = CATEGORY_COLORS.get(row['category'], 'FFFFFF')
        row_data = [
            row['contact_name'],
            row['category'],
            row['priority'],
            row['days_since'],
            row['phone'] or '—',
            row['last_message'],
            row['suggested'],
            row['folder'],
        ]
        for c_idx, val in enumerate(row_data, 1):
            c = ws1.cell(row=r_idx, column=c_idx, value=val)
            c.border = cell_border()
            c.alignment = Alignment(vertical='center', wrap_text=True)
            c.font = Font(size=9, name='Arial')
            if c_idx <= 4:
                c.fill = PatternFill('solid', fgColor=color)
        ws1.row_dimensions[r_idx].height = 32

    # ── Tab 2: Priority Call List (priority 7+, has phone) ──
    ws2 = wb.create_sheet("📞 Priority Call List")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = 'A3'
    ws2.merge_cells('A1:F1')
    t2 = ws2['A1']
    t2.value = "Priority Call List — Contacts with Phone Numbers (Priority 7+)"
    t2.font = Font(bold=True, size=13, color='FFFFFF', name='Arial')
    t2.fill = PatternFill('solid', fgColor='E94560')
    t2.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 36
    header_style(ws2, 2, ['Contact Name', 'Category', 'Priority', 'Phone', 'Days Since', 'Suggested Outreach'], fill_color='E94560')
    for i, w in enumerate([22, 28, 10, 16, 14, 50], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    priority_calls = [r for r in classified if r['priority'] >= 7 and r['phone']]
    for r_idx, row in enumerate(priority_calls, 3):
        data = [row['contact_name'], row['category'], row['priority'], row['phone'], row['days_since'], row['suggested']]
        for c_idx, val in enumerate(data, 1):
            c = ws2.cell(row=r_idx, column=c_idx, value=val)
            c.border = cell_border()
            c.alignment = Alignment(vertical='center', wrap_text=True)
            c.font = Font(size=9, name='Arial')
            if r_idx % 2 == 0:
                c.fill = PatternFill('solid', fgColor='FFF0F3')
        ws2.row_dimensions[r_idx].height = 28

    # ── Tab 3: Needs Reply ──
    ws3 = wb.create_sheet("💬 Needs Reply")
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = 'A3'
    ws3.merge_cells('A1:E1')
    t3 = ws3['A1']
    t3.value = "Amy Needs to Reply — They Reached Out First"
    t3.font = Font(bold=True, size=13, color='FFFFFF', name='Arial')
    t3.fill = PatternFill('solid', fgColor='FF6B35')
    t3.alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 36
    header_style(ws3, 2, ['Contact Name', 'Days Since', 'Phone', 'Last Message Preview', 'Folder'], fill_color='FF6B35')
    for i, w in enumerate([22, 14, 16, 50, 28], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    needs_reply = [r for r in classified if 'Amy Needs to Reply' in r['category']]
    for r_idx, row in enumerate(needs_reply, 3):
        data = [row['contact_name'], row['days_since'], row['phone'] or '—', row['last_message'], row['folder']]
        for c_idx, val in enumerate(data, 1):
            c = ws3.cell(row=r_idx, column=c_idx, value=val)
            c.border = cell_border()
            c.alignment = Alignment(vertical='center', wrap_text=True)
            c.font = Font(size=9, name='Arial')
        ws3.row_dimensions[r_idx].height = 28

    # ── Tab 4: Past Customers ──
    ws4 = wb.create_sheet("🏆 Past Customers")
    ws4.sheet_view.showGridLines = False
    ws4.freeze_panes = 'A3'
    ws4.merge_cells('A1:E1')
    t4 = ws4['A1']
    t4.value = "Past Customers — Trade-Up Ready & Check-In"
    t4.font = Font(bold=True, size=13, color='FFFFFF', name='Arial')
    t4.fill = PatternFill('solid', fgColor='0F3460')
    t4.alignment = Alignment(horizontal='center', vertical='center')
    ws4.row_dimensions[1].height = 36
    header_style(ws4, 2, ['Contact Name', 'Category', 'Days Since', 'Phone', 'Suggested Outreach'], fill_color='0F3460')
    for i, w in enumerate([22, 28, 14, 16, 50], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    past = [r for r in classified if 'Past Customer' in r['category'] or 'Recent Customer' in r['category']]
    for r_idx, row in enumerate(past, 3):
        data = [row['contact_name'], row['category'], row['days_since'], row['phone'] or '—', row['suggested']]
        for c_idx, val in enumerate(data, 1):
            c = ws4.cell(row=r_idx, column=c_idx, value=val)
            c.border = cell_border()
            c.alignment = Alignment(vertical='center', wrap_text=True)
            c.font = Font(size=9, name='Arial')
        ws4.row_dimensions[r_idx].height = 28

    # ── Tab 5: Summary ──
    ws5 = wb.create_sheet("📊 Summary")
    ws5.sheet_view.showGridLines = False
    ws5.column_dimensions['A'].width = 35
    ws5.column_dimensions['B'].width = 15

    ws5.merge_cells('A1:B1')
    t5 = ws5['A1']
    t5.value = "Rekindl — Opportunity Summary"
    t5.font = Font(bold=True, size=14, color='FFFFFF', name='Arial')
    t5.fill = PatternFill('solid', fgColor='1A1A2E')
    t5.alignment = Alignment(horizontal='center', vertical='center')
    ws5.row_dimensions[1].height = 36

    cat_counts = Counter(r['category'] for r in classified)
    for r_idx, cat in enumerate(CATEGORY_ORDER, 3):
        count = cat_counts.get(cat, 0)
        color = CATEGORY_COLORS.get(cat, 'FFFFFF')
        c1 = ws5.cell(row=r_idx, column=1, value=cat)
        c2 = ws5.cell(row=r_idx, column=2, value=count)
        for c in [c1, c2]:
            c.fill = PatternFill('solid', fgColor=color)
            c.border = cell_border()
            c.alignment = Alignment(vertical='center')
            c.font = Font(size=10, name='Arial', bold=(c == c2))
        ws5.row_dimensions[r_idx].height = 24

    # Total
    total_row = 3 + len(CATEGORY_ORDER)
    c_tot1 = ws5.cell(row=total_row, column=1, value="TOTAL")
    c_tot2 = ws5.cell(row=total_row, column=2, value=len(classified))
    for c in [c_tot1, c_tot2]:
        c.fill = PatternFill('solid', fgColor='1A1A2E')
        c.font = Font(bold=True, color='FFFFFF', size=11, name='Arial')
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws5.row_dimensions[total_row].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ─────────────────────────────────────────
# STEP 5 — GENERATE GHL CSV
# ─────────────────────────────────────────

def build_ghl_csv(classified, min_priority=7):
    priority_contacts = [r for r in classified if r['priority'] >= min_priority and r['contact_name'].lower() not in ['facebook user', 'unknown', '']]
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=['First Name', 'Last Name', 'Phone', 'Tags', 'Source', 'Category', 'Days Since Contact', 'Priority'])
    writer.writeheader()
    for r in priority_contacts:
        parts = r['contact_name'].split(' ', 1)
        first = parts[0]
        last = parts[1] if len(parts) > 1 else ''
        cat_tag = r['category'].lower().replace(' ', '-').replace('–', '').replace('—', '').replace('  ', '-')
        tags = f"rekindl,{cat_tag}"
        writer.writerow({
            'First Name': first,
            'Last Name': last,
            'Phone': r['phone'] or '',
            'Tags': tags,
            'Source': 'Facebook Messenger',
            'Category': r['category'],
            'Days Since Contact': r['days_since'],
            'Priority': r['priority'],
        })
    return buf.getvalue(), len(priority_contacts)

# ─────────────────────────────────────────
# MAIN UI
# ─────────────────────────────────────────

col1, col2 = st.columns([2, 1])
with col1:
    st.markdown("### 📁 Upload Facebook Export")
    uploaded_file = st.file_uploader(
        "Upload your Facebook Messenger export (.zip)",
        type=['zip'],
        help="Download from Facebook: Settings → Your Information → Download Your Information → Messages"
    )
with col2:
    st.markdown("### 👤 Salesperson Name")
    sender_name = st.text_input(
        "Enter the name exactly as it appears in the messages",
        placeholder="e.g. Amy Gauthier",
        help="This must match the sender_name field in your Facebook export exactly."
    )
    min_priority = st.slider("Min priority for GHL export", 1, 10, 7, help="Only contacts with this priority or higher will be included in the GHL CSV export.")

if uploaded_file and sender_name:
    if st.button("🔥 Run Rekindl Analysis", use_container_width=True, type="primary"):
        with st.spinner("Processing your data... this may take a minute for large exports."):
            try:
                zip_bytes = uploaded_file.read()

                # Load conversations
                progress = st.progress(0, text="Loading conversations...")
                conversations = load_conversations_from_zip(zip_bytes, sender_name)
                progress.progress(33, text=f"Loaded {len(conversations)} conversations. Running brand voice analysis...")

                # Brand voice
                all_my_msgs = []
                for c in conversations:
                    all_my_msgs.extend(c['my_messages'])
                voice = analyse_voice(all_my_msgs)
                progress.progress(66, text="Classifying opportunities...")

                # Classify
                classified = classify_all(conversations)
                progress.progress(90, text="Building outputs...")

                # Build downloads
                excel_bytes = build_excel(classified, sender_name)
                csv_str, ghl_count = build_ghl_csv(classified, min_priority)
                progress.progress(100, text="Done!")

                st.session_state['done'] = True
                st.session_state['voice'] = voice
                st.session_state['classified'] = classified
                st.session_state['excel_bytes'] = excel_bytes
                st.session_state['csv_str'] = csv_str
                st.session_state['ghl_count'] = ghl_count
                st.session_state['sender_name'] = sender_name
                st.session_state['total_convos'] = len(conversations)

            except Exception as e:
                st.error(f"Something went wrong: {e}")
                st.exception(e)

# ─────────────────────────────────────────
# RESULTS
# ─────────────────────────────────────────

if st.session_state.get('done'):
    voice      = st.session_state['voice']
    classified = st.session_state['classified']
    excel_bytes= st.session_state['excel_bytes']
    csv_str    = st.session_state['csv_str']
    ghl_count  = st.session_state['ghl_count']
    name       = st.session_state['sender_name']
    total      = st.session_state['total_convos']

    st.success(f"✅ Analysis complete for **{name}** — {total} conversations processed.")
    st.divider()

    # ── Downloads ──
    st.markdown("### 📥 Downloads")
    dc1, dc2 = st.columns(2)
    with dc1:
        st.download_button(
            label="⬇️ Download Sales Opportunity Tracker (.xlsx)",
            data=excel_bytes,
            file_name=f"Rekindl_Opportunities_{name.replace(' ','_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    with dc2:
        st.download_button(
            label=f"⬇️ Download GHL Import CSV ({ghl_count} contacts)",
            data=csv_str,
            file_name=f"Rekindl_GHL_Import_{name.replace(' ','_')}.csv",
            mime="text/csv",
            use_container_width=True
        )

    st.divider()

    # ── Tabs ──
    tab1, tab2, tab3 = st.tabs(["📊 Opportunity Summary", "🎤 Brand Voice", "📋 Full Lead List"])

    with tab1:
        cat_counts = Counter(r['category'] for r in classified)
        st.markdown(f"**{len(classified)} total contacts classified across {len(CATEGORY_ORDER)} categories**")
        st.markdown("")

        # Top metrics
        m1, m2, m3, m4 = st.columns(4)
        pc = sum(1 for r in classified if 'Past Customer' in r['category'])
        hl = sum(1 for r in classified if 'Hot Lead' in r['category'])
        nr = sum(1 for r in classified if 'Amy Needs to Reply' in r['category'])
        ph = sum(1 for r in classified if r['phone'])

        with m1:
            st.markdown(f'<div class="metric-card"><h3>{pc}</h3><p>Past Customers</p></div>', unsafe_allow_html=True)
        with m2:
            st.markdown(f'<div class="metric-card"><h3>{hl}</h3><p>Hot Leads</p></div>', unsafe_allow_html=True)
        with m3:
            st.markdown(f'<div class="metric-card"><h3>{nr}</h3><p>Need Reply Now</p></div>', unsafe_allow_html=True)
        with m4:
            st.markdown(f'<div class="metric-card"><h3>{ph}</h3><p>Have Phone Numbers</p></div>', unsafe_allow_html=True)

        st.markdown("")
        # Category breakdown
        for cat in CATEGORY_ORDER:
            count = cat_counts.get(cat, 0)
            if count > 0:
                pct = count / len(classified) * 100
                st.markdown(f"**{cat}** — {count} contacts")
                st.progress(pct / 100)

    with tab2:
        if voice:
            v1, v2 = st.columns(2)
            with v1:
                st.markdown(f"**Total messages analysed:** {voice['total_messages']:,}")
                st.markdown(f"**Average message length:** {voice['avg_message_length']} words")
                st.markdown(f"**Messages with emoji:** {voice['emoji_pct']}%")
                st.markdown(f"**Messages with no end punctuation:** {voice['no_punct_pct']}%")

                st.markdown("#### Top Openers")
                for opener, count in voice['openers']:
                    st.markdown(f"- **\"{opener.title()}\"** — used {count} times")

                st.markdown("#### Signature Phrases (Trigrams)")
                for phrase, count in voice['top_trigrams'][:8]:
                    st.markdown(f'<div class="voice-quote">"{phrase}" <span style="color:#999;font-size:0.8rem">({count}×)</span></div>', unsafe_allow_html=True)

            with v2:
                st.markdown("#### Tone Profile")
                tone_sorted = sorted(voice['tone_scores'].items(), key=lambda x: -x[1])
                top_score = tone_sorted[0][1] if tone_sorted else 1
                for tone, score in tone_sorted:
                    if score > 0:
                        pct = score / top_score
                        st.markdown(f"**{tone}**")
                        st.progress(pct)

                st.markdown("#### Top Words")
                chips_html = ''.join(f'<span class="tag-chip">{w}</span>' for w, _ in voice['top_words'][:15])
                st.markdown(chips_html, unsafe_allow_html=True)

                st.markdown("#### Top Phrases (Bigrams)")
                for phrase, count in voice['top_bigrams'][:8]:
                    st.markdown(f'<div class="voice-quote">"{phrase}" ({count}×)</div>', unsafe_allow_html=True)

    with tab3:
        st.markdown(f"**Showing top {min(100, len(classified))} of {len(classified)} contacts**")
        import pandas as pd
        df = pd.DataFrame([{
            'Name': r['contact_name'],
            'Category': r['category'],
            'Priority': r['priority'],
            'Days Since': r['days_since'],
            'Phone': r['phone'] or '—',
            'Last Message': r['last_message'][:80] + '...' if len(r['last_message']) > 80 else r['last_message'],
        } for r in classified[:100]])
        st.dataframe(df, use_container_width=True, hide_index=True)

else:
    st.info("👆 Upload a Facebook Messenger export zip file and enter the salesperson's name to get started.")
    st.markdown("""
    **How to export your Facebook messages:**
    1. Go to Facebook → Settings & Privacy → Settings
    2. Click **Your Facebook Information** → **Download Your Information**
    3. Select **Messages** only, set format to **JSON**, quality to **Low**
    4. Request the download, wait for the email, then download the zip
    5. Upload that zip file here
    """)
